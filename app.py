# -*- coding: utf-8 -*-
"""
NailVesta · 深深达名单生成器
-------------------------------------------------
输入两份文件：
  1) 深度达人 List（既有名单，键 = handle）
  2) TikTok Creator 数据（某时段，键 = Creator username，含 GMV / orders / items 等）

逻辑：
  把两份名单按清洗后的 handle 比对，过滤掉这段时间「已经没在出单」的深度达人，
  保留仍在持续出单的，组成一份更高级的「深深达」名单。

出单门槛（GMV / 出单数 / 件数 + 最小值）可在左侧侧边栏自行调整。
"""

import io
import re
import sys
import subprocess
import importlib

# --------------------------------------------------------------------------- #
#  自我修复：若部署环境缺 openpyxl（pandas 读 .xlsx 的引擎），启动时自动安装。
#  这样即使 Streamlit Cloud 没读到 requirements.txt 也能正常运行。
# --------------------------------------------------------------------------- #
def _ensure(pkg, import_name=None):
    try:
        importlib.import_module(import_name or pkg)
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "--quiet", pkg], check=True)
        importlib.invalidate_caches()

for _p in ("openpyxl", "pandas"):
    _ensure(_p)

import pandas as pd
import streamlit as st

st.set_page_config(page_title="NailVesta · 深深达名单生成器", page_icon="💅", layout="wide")

# ----------------------------------------------------------------------------- #
#  Handle 清洗（与既有 Streamlit 管线一致：去全/半形括号、去 @、去结尾 -数字、转小写）
# ----------------------------------------------------------------------------- #
def clean_handle(h, strip_paren=True, strip_suffix=True):
    if pd.isna(h):
        return None
    s = str(h).strip().lower()
    if strip_paren:
        s = re.sub(r"（.*?）", "", s)   # 全形括号，例如 inessaakin（改名inessaak）
        s = re.sub(r"\(.*?\)", "", s)   # 半形括号
    s = s.lstrip("@")
    if strip_suffix:
        s = re.sub(r"-\d+$", "", s)     # 结尾 -数字，例如 gicellex-2
    return s.strip() or None


def to_num(series):
    """把 '--'、千分位、空字串等转成数字。"""
    return pd.to_numeric(
        series.astype(str).str.replace(",", "", regex=False).replace({"--": None, "": None, "nan": None}),
        errors="coerce",
    ).fillna(0)


def first_non_null(s):
    s = s.dropna()
    return s.iloc[0] if len(s) else None


@st.cache_data(show_spinner=False)
def read_excel(file_bytes):
    return pd.read_excel(io.BytesIO(file_bytes))


# 候选键名（自动侦测用）
DEEP_KEY_CANDIDATES = ["handle", "Handle", "Creator username", "username", "用户名"]
TT_KEY_CANDIDATES = ["Creator username", "creator username", "username", "handle", "用户名"]
GMV_CANDIDATES = ["Affiliate GMV", "GMV", "affiliate gmv"]
ORDER_CANDIDATES = ["Affiliate orders", "Orders", "affiliate orders"]
ITEM_CANDIDATES = ["Items sold", "items sold", "Affiliate items sold"]
REFUND_GMV_CANDIDATES = ["Affiliate refunded GMV", "Refunded GMV"]
FOLLOWERS_CANDIDATES = ["Affiliate followers", "Followers"]
AOV_CANDIDATES = ["Avg. order value", "AOV", "Average order value"]

# 深度名单里想带进结果的业务字段（刻意不带 人種 等敏感栏位）
DEEP_CARRY = ["Level", "评级", "深度合作Status", "深度前出单数", "终止合作",
              "名字", "Email", "Phone", "Tiktok Link", "aff link"]


def pick(cols, candidates):
    for c in candidates:
        if c in cols:
            return c
    return None


# ----------------------------------------------------------------------------- #
#  侧边栏：上传 + 设定
# ----------------------------------------------------------------------------- #
st.sidebar.header("① 上传文件")
deep_file = st.sidebar.file_uploader("深度达人 List (.xlsx)", type=["xlsx", "xls"])
tt_file = st.sidebar.file_uploader("TikTok Creator 数据 (.xlsx)", type=["xlsx", "xls"])

st.sidebar.header("② Handle 清洗")
strip_paren = st.sidebar.checkbox("去除括号内容 （改名…）", value=True)
strip_suffix = st.sidebar.checkbox("去除结尾 -数字 后缀（gicellex-2 → gicellex）", value=True)

st.sidebar.header("③ 出单门槛")
metric_label = st.sidebar.selectbox(
    "判定「仍在出单」的指标",
    ["出单数 (Affiliate orders)", "GMV (Affiliate GMV)", "件数 (Items sold)"],
    index=0,
)
metric_key = {"出单数": "orders", "GMV": "gmv", "件数": "items"}[metric_label.split(" ")[0]]

if metric_key == "gmv":
    min_val = st.sidebar.number_input("最小 GMV（含）", min_value=0.0, value=1.0, step=10.0)
elif metric_key == "items":
    min_val = st.sidebar.number_input("最小件数（含）", min_value=0, value=1, step=1)
else:
    min_val = st.sidebar.number_input("最小出单数（含）", min_value=0, value=1, step=1)

exclude_terminated = st.sidebar.checkbox("排除已标记『终止合作』的达人", value=False)

st.title("💅 NailVesta · 深深达名单生成器")
st.caption("过滤掉这段时间已经没在出单的深度达人，产出一份更精炼的深深达名单。")

if not deep_file or not tt_file:
    st.info("请在左侧分别上传 **深度达人 List** 与 **TikTok Creator 数据** 两份文件。")
    st.stop()

# ----------------------------------------------------------------------------- #
#  读取
# ----------------------------------------------------------------------------- #
deep_raw = read_excel(deep_file.getvalue())
tt_raw = read_excel(tt_file.getvalue())

deep_key = pick(deep_raw.columns, DEEP_KEY_CANDIDATES)
tt_key = pick(tt_raw.columns, TT_KEY_CANDIDATES)

with st.sidebar.expander("⚙️ 键名/指标对应（自动侦测，可手动改）", expanded=False):
    deep_key = st.selectbox("深度名单 handle 栏", list(deep_raw.columns),
                            index=list(deep_raw.columns).index(deep_key) if deep_key else 0)
    tt_key = st.selectbox("TikTok username 栏", list(tt_raw.columns),
                          index=list(tt_raw.columns).index(tt_key) if tt_key else 0)

gmv_col = pick(tt_raw.columns, GMV_CANDIDATES)
order_col = pick(tt_raw.columns, ORDER_CANDIDATES)
item_col = pick(tt_raw.columns, ITEM_CANDIDATES)
refund_col = pick(tt_raw.columns, REFUND_GMV_CANDIDATES)
fol_col = pick(tt_raw.columns, FOLLOWERS_CANDIDATES)
aov_col = pick(tt_raw.columns, AOV_CANDIDATES)

# ----------------------------------------------------------------------------- #
#  清洗 + 去重（深度名单：每位达人取一条，带最有信息量的业务字段）
# ----------------------------------------------------------------------------- #
deep = deep_raw.copy()
deep["__key"] = deep[deep_key].map(lambda x: clean_handle(x, strip_paren, strip_suffix))
deep = deep[deep["__key"].notna()]

agg = {}
for c in DEEP_CARRY:
    if c in deep.columns:
        agg[c] = "max" if c == "深度前出单数" else first_non_null
deep_u = deep.groupby("__key").agg(agg).reset_index()
deep_u = deep_u.rename(columns={"__key": "handle"})

# ----------------------------------------------------------------------------- #
#  清洗 + 聚合（TikTok：同 username 多行合并）
# ----------------------------------------------------------------------------- #
tt = tt_raw.copy()
tt["__key"] = tt[tt_key].map(lambda x: clean_handle(x, strip_paren, strip_suffix))
tt = tt[tt["__key"].notna()]

tt["gmv"] = to_num(tt[gmv_col]) if gmv_col else 0
tt["orders"] = to_num(tt[order_col]) if order_col else 0
tt["items"] = to_num(tt[item_col]) if item_col else 0
tt["refund_gmv"] = to_num(tt[refund_col]) if refund_col else 0
tt["followers"] = to_num(tt[fol_col]) if fol_col else 0

tt_g = tt.groupby("__key").agg(
    gmv=("gmv", "sum"), orders=("orders", "sum"), items=("items", "sum"),
    refund_gmv=("refund_gmv", "sum"), followers=("followers", "max"),
).reset_index().rename(columns={"__key": "handle"})

# ----------------------------------------------------------------------------- #
#  比对
# ----------------------------------------------------------------------------- #
merged = deep_u.merge(tt_g, on="handle", how="left")
for c in ["gmv", "orders", "items", "refund_gmv", "followers"]:
    merged[c] = merged[c].fillna(0)
merged["in_period"] = merged["handle"].isin(set(tt_g["handle"]))
merged["退款率"] = (merged["refund_gmv"] / merged["gmv"]).where(merged["gmv"] > 0, 0).round(3)

# 门槛
metric_series = merged[metric_key]
qualifies = metric_series >= min_val
if exclude_terminated and "终止合作" in merged.columns:
    terminated = pd.to_numeric(merged["终止合作"], errors="coerce").fillna(0) > 0
    qualifies = qualifies & ~terminated
merged["深深达"] = qualifies

kept = merged[merged["深深达"]].sort_values("gmv", ascending=False).reset_index(drop=True)
dropped = merged[~merged["深深达"]].sort_values("gmv", ascending=False).reset_index(drop=True)

# ----------------------------------------------------------------------------- #
#  概览
# ----------------------------------------------------------------------------- #
c1, c2, c3, c4 = st.columns(4)
c1.metric("深度达人总数", f"{len(merged):,}")
c2.metric("时段内有数据", f"{int(merged['in_period'].sum()):,}")
c3.metric("✅ 深深达（保留）", f"{len(kept):,}")
c4.metric("❌ 已剔除", f"{len(dropped):,}")

c5, c6, c7 = st.columns(3)
c5.metric("深深达 总 GMV", f"${kept['gmv'].sum():,.0f}")
c6.metric("深深达 总出单数", f"{int(kept['orders'].sum()):,}")
c7.metric("深深达 平均 GMV", f"${kept['gmv'].mean():,.0f}" if len(kept) else "$0")

st.caption(
    f"门槛：**{metric_label} ≥ {min_val}**"
    + ("，且排除『终止合作』" if exclude_terminated else "")
    + f"。未出现在 TikTok 数据中的 {int((~merged['in_period']).sum())} 位达人视为未出单、一律剔除。"
)

# ----------------------------------------------------------------------------- #
#  展示用栏位
# ----------------------------------------------------------------------------- #
display_cols = ["handle", "Level", "深度前出单数", "gmv", "orders", "items",
                "退款率", "followers", "名字", "Email", "Phone", "Tiktok Link", "aff link"]
display_cols = [c for c in display_cols if c in kept.columns]
rename_map = {"gmv": "时段GMV", "orders": "时段出单数", "items": "时段件数", "followers": "粉丝数"}

tab1, tab2, tab3 = st.tabs(["✅ 深深达名单", "❌ 已剔除", "📊 GMV 分布"])

with tab1:
    st.dataframe(kept[display_cols].rename(columns=rename_map),
                 use_container_width=True, hide_index=True)

with tab2:
    reason = pd.Series("低于门槛", index=dropped.index)
    reason[~dropped["in_period"]] = "时段内无数据/未出单"
    if exclude_terminated and "终止合作" in dropped.columns:
        reason[pd.to_numeric(dropped["终止合作"], errors="coerce").fillna(0) > 0] = "已终止合作"
    d = dropped[display_cols].rename(columns=rename_map).copy()
    d.insert(1, "剔除原因", reason.values)
    st.dataframe(d, use_container_width=True, hide_index=True)

with tab3:
    pos = kept[kept["gmv"] > 0]
    if len(pos):
        st.bar_chart(pos.set_index("handle")["gmv"].head(30))
        st.caption("Top 30 深深达（按时段 GMV）")
    else:
        st.write("没有 GMV > 0 的达人。")

# ----------------------------------------------------------------------------- #
#  导出 Excel
# ----------------------------------------------------------------------------- #
def build_excel(kept_df, dropped_df, full_df):
    out = io.BytesIO()
    cols = display_cols
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        kept_df[cols].rename(columns=rename_map).to_excel(xw, sheet_name="深深达名单", index=False)
        dd = dropped_df[cols].rename(columns=rename_map)
        dd.to_excel(xw, sheet_name="已剔除", index=False)
        full_df.to_excel(xw, sheet_name="全部比对", index=False)
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        hdr_fill = PatternFill("solid", fgColor="4472C4")
        hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ws in xw.book.worksheets:
            for cell in ws[1]:
                cell.fill, cell.font = hdr_fill, hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            ws.freeze_panes = "A2"
            for col in ws.columns:
                width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 42)
    return out.getvalue()

st.sidebar.header("④ 导出")
st.sidebar.download_button(
    "📥 下载结果 Excel",
    data=build_excel(kept, dropped, merged.drop(columns=["深深达"], errors="ignore")),
    file_name="NailVesta_深深达名单.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
